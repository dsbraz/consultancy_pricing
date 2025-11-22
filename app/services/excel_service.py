from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from sqlalchemy.orm import Session

from app.models.models import Project
from app.services.pricing_service import PricingService
from app.services.calendar_service import CalendarService


class ExcelExportService:
    def __init__(self, db: Session):
        self.db = db
        self.pricing_service = PricingService(db)
        self.calendar_service = CalendarService(country_code='BR')

    def export_project_to_excel(self, project: Project) -> BytesIO:
        """
        Exporta um projeto completo para arquivo Excel com 3 abas:
        1. Informações do Projeto
        2. Resumo Financeiro
        3. Tabela de Alocação
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        self._create_project_info_sheet(wb, project)
        self._create_financial_summary_sheet(wb, project)
        self._create_allocation_table_sheet(wb, project)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _create_project_info_sheet(self, wb: Workbook, project: Project):
        """Cria a aba de Informações do Projeto"""
        ws = wb.create_sheet("Informações do Projeto")
        
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        data = [
            ["Campo", "Valor"],
            ["Nome do Projeto", project.name],
            ["Data de Início", project.start_date.strftime("%d/%m/%Y")],
            ["Duração", f"{project.duration_months} meses"],
            ["Taxa de Impostos", f"{project.tax_rate}%"],
            ["Taxa de Margem", f"{project.margin_rate}%"],
        ]
        
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def _create_financial_summary_sheet(self, wb: Workbook, project: Project):
        """Cria a aba de Resumo Financeiro"""
        ws = wb.create_sheet("Resumo Financeiro")
        
        pricing = self.pricing_service.calculate_project_pricing(project)
        
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        data = [
            ["Métrica", "Valor"],
            ["Custo Total", f"R$ {pricing['total_cost']:,.2f}"],
            ["Venda Total", f"R$ {pricing['total_selling']:,.2f}"],
            ["Margem Total", f"R$ {pricing['total_margin']:,.2f}"],
            ["Impostos Totais", f"R$ {pricing['total_tax']:,.2f}"],
            ["Preço Final", f"R$ {pricing['final_price']:,.2f}"],
            ["Margem Final (%)", f"{pricing['final_margin_percent']:.2f}%"],
        ]
        
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25

    def _create_allocation_table_sheet(self, wb: Workbook, project: Project):
        """Cria a aba de Tabela de Alocação"""
        ws = wb.create_sheet("Tabela de Alocação")
        
        weeks = self.calendar_service.get_weekly_breakdown(
            project.start_date, 
            project.duration_months
        )
        
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        
        headers = [
            "ID", "Nome", "Função", "Nível", 
            "Custo Horário", "Taxa de Venda"
        ]
        
        for week in weeks:
            from datetime import datetime as dt
            week_start_date = dt.fromisoformat(week['week_start']).date() if isinstance(week['week_start'], str) else week['week_start']
            week_label = f"Semana {week['week_number']}\n{week_start_date.strftime('%d/%m/%Y')}\n({week['available_hours']}h disponíveis)"
            headers.append(week_label)
        
        headers.append("Total de Horas")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        row_idx = 2
        for allocation in project.allocations:
            professional = allocation.professional
            
            weekly_hours = {}
            for weekly_alloc in allocation.weekly_allocations:
                weekly_hours[weekly_alloc.week_number] = {
                    'allocated': weekly_alloc.hours_allocated,
                    'available': weekly_alloc.available_hours
                }
            
            row_data = [
                professional.pid,
                professional.name,
                professional.role,
                professional.level,
                f"R$ {professional.hourly_cost:.2f}",
                f"R$ {allocation.selling_hourly_rate:.2f}",
            ]
            
            total_hours = 0
            for week in weeks:
                week_num = week['week_number']
                if week_num in weekly_hours:
                    allocated = weekly_hours[week_num]['allocated']
                    row_data.append(allocated if allocated > 0 else "")
                    total_hours += allocated
                else:
                    row_data.append("")
            
            row_data.append(f"{total_hours:.1f}")
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row_idx += 1
        
        ws.column_dimensions['A'].width = 12  # ID
        ws.column_dimensions['B'].width = 25  # Nome
        ws.column_dimensions['C'].width = 20  # Função
        ws.column_dimensions['D'].width = 15  # Nível
        ws.column_dimensions['E'].width = 15  # Custo Horário
        ws.column_dimensions['F'].width = 15  # Taxa de Venda
        
        for i in range(len(weeks)):
            col_letter = get_column_letter(7 + i)
            ws.column_dimensions[col_letter].width = 15
        
        total_col = get_column_letter(7 + len(weeks))
        ws.column_dimensions[total_col].width = 15
        
        ws.row_dimensions[1].height = 45
