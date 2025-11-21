import pytest
from fastapi.testclient import TestClient
from io import BytesIO
from openpyxl import load_workbook
from datetime import date

from app.main import app
from app.database import get_db, Base, engine
from app.models.models import Project, Professional, ProjectAllocation, WeeklyAllocation
from app.services.calendar_service import CalendarService


# Test client
client = TestClient(app)


@pytest.fixture(scope="function", autouse=True)
def setup_database():
    """Setup test database before each test and cleanup after"""
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


def test_export_project_excel():
    """Test exporting a project to Excel format"""
    
    # Create a professional
    prof_data = {
        "pid": "TEST-001",
        "name": "João Silva",
        "role": "Desenvolvedor",
        "level": "Senior",
        "is_vacancy": False,
        "hourly_cost": 100.0
    }
    prof_response = client.post("/professionals/", json=prof_data)
    assert prof_response.status_code == 200
    professional = prof_response.json()
    
    # Create a project
    project_data = {
        "name": "Projeto Teste Excel",
        "start_date": "2025-01-01",
        "duration_months": 2,
        "tax_rate": 11.0,
        "margin_rate": 40.0,
        "allocations": []
    }
    proj_response = client.post("/projects/", json=project_data)
    assert proj_response.status_code == 200
    project = proj_response.json()
    project_id = project["id"]
    
    # Add professional to project
    add_prof_response = client.post(
        f"/projects/{project_id}/allocations/?professional_id={professional['id']}"
    )
    assert add_prof_response.status_code == 200
    
    # Get allocation table to allocate hours
    alloc_table = client.get(f"/projects/{project_id}/allocation_table")
    assert alloc_table.status_code == 200
    alloc_data = alloc_table.json()
    
    # Update some weekly hours
    updates = []
    if len(alloc_data["allocations"]) > 0:
        allocation = alloc_data["allocations"][0]
        weekly_hours = allocation["weekly_hours"]
        
        # Set hours for first 2 weeks (respecting available hours)
        week_numbers = sorted(weekly_hours.keys(), key=int)[:2]
        for week_num_str in week_numbers:
            week_num = int(week_num_str)
            weekly_alloc = weekly_hours[week_num_str]
            available = weekly_alloc["available_hours"]
            # Allocate 80% of available hours, or max 40
            hours_to_allocate = min(available * 0.8, 40.0)
            updates.append({
                "weekly_allocation_id": weekly_alloc["id"],
                "hours_allocated": hours_to_allocate
            })
    
    if updates:
        update_response = client.put(f"/projects/{project_id}/allocations", json=updates)
        assert update_response.status_code == 200
    
    # Export to Excel
    export_response = client.get(f"/projects/{project_id}/export_excel")
    assert export_response.status_code == 200
    assert export_response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in export_response.headers["content-disposition"]
    
    # Load and verify Excel content
    excel_bytes = BytesIO(export_response.content)
    workbook = load_workbook(excel_bytes)
    
    # Verify sheet names
    sheet_names = workbook.sheetnames
    assert "Informações do Projeto" in sheet_names
    assert "Resumo Financeiro" in sheet_names
    assert "Tabela de Alocação" in sheet_names
    
    # Verify Project Info Sheet
    info_sheet = workbook["Informações do Projeto"]
    assert info_sheet["A2"].value == "Nome do Projeto"
    assert info_sheet["B2"].value == "Projeto Teste Excel"
    assert info_sheet["A3"].value == "Data de Início"
    assert info_sheet["A4"].value == "Duração"
    assert "2 meses" in info_sheet["B4"].value
    assert info_sheet["A5"].value == "Taxa de Impostos"
    assert "11" in str(info_sheet["B5"].value)
    assert info_sheet["A6"].value == "Taxa de Margem"
    assert "40" in str(info_sheet["B6"].value)
    
    # Verify Financial Summary Sheet
    financial_sheet = workbook["Resumo Financeiro"]
    assert financial_sheet["A1"].value == "Métrica"
    assert financial_sheet["B1"].value == "Valor"
    assert financial_sheet["A2"].value == "Custo Total"
    assert financial_sheet["A3"].value == "Venda Total"
    assert financial_sheet["A4"].value == "Margem Total"
    assert financial_sheet["A5"].value == "Impostos Totais"
    assert financial_sheet["A6"].value == "Preço Final"
    assert financial_sheet["A7"].value == "Margem Final (%)"
    
    # Verify Allocation Table Sheet
    allocation_sheet = workbook["Tabela de Alocação"]
    assert allocation_sheet["A1"].value == "ID"
    assert allocation_sheet["B1"].value == "Nome"
    assert allocation_sheet["C1"].value == "Função"
    assert allocation_sheet["D1"].value == "Nível"
    assert allocation_sheet["E1"].value == "Custo Horário"
    assert allocation_sheet["F1"].value == "Taxa de Venda"
    
    # Verify professional data in allocation table
    assert allocation_sheet["A2"].value == "TEST-001"  # PID
    assert allocation_sheet["B2"].value == "João Silva"  # Name
    assert allocation_sheet["C2"].value == "Desenvolvedor"  # Role
    assert allocation_sheet["D2"].value == "Senior"  # Level
    assert "100" in str(allocation_sheet["E2"].value)  # Cost
    
    print("✅ All Excel export tests passed!")


def test_export_nonexistent_project():
    """Test exporting a non-existent project returns 404"""
    export_response = client.get("/projects/99999/export_excel")
    assert export_response.status_code == 404
    assert "não encontrado" in export_response.json()["detail"].lower()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
