"""
Verification script for Excel export functionality.
Tests exporting a project to Excel format against running server.
"""
import requests
import sys
from io import BytesIO
from openpyxl import load_workbook
import datetime

BASE_URL = "http://localhost:8080"

def verify_excel_export():
    print("Starting Excel Export Verification...")
    
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    
    # Create a professional
    print("\n1. Creating professional...")
    prof_data = {
        "pid": f"TEST-{timestamp}",
        "name": "João Silva",
        "role": "Desenvolvedor",
        "level": "Senior",
        "is_template": False,
        "hourly_cost": 100.0
    }
    prof_response = requests.post(f"{BASE_URL}/professionals/", json=prof_data)
    if prof_response.status_code != 200:
        print(f"❌ Failed to create professional: {prof_response.text}")
        sys.exit(1)
    professional = prof_response.json()
    print(f"✅ Professional created: {professional['name']}")
    
    # Create a project
    print("\n2. Creating project...")
    project_data = {
        "name": f"Projeto Teste Excel {timestamp}",
        "start_date": "2025-01-01",
        "duration_months": 2,
        "tax_rate": 11.0,
        "margin_rate": 40.0
    }
    proj_response = requests.post(f"{BASE_URL}/projects/", json=project_data)
    if proj_response.status_code != 200:
        print(f"❌ Failed to create project: {proj_response.text}")
        sys.exit(1)
    project = proj_response.json()
    project_id = project["id"]
    print(f"✅ Project created: {project['name']}")
    
    # Add professional to project
    print("\n3. Adding professional to project...")
    add_prof_response = requests.post(
        f"{BASE_URL}/projects/{project_id}/allocations/?professional_id={professional['id']}"
    )
    if add_prof_response.status_code != 200:
        print(f"❌ Failed to add professional: {add_prof_response.text}")
        sys.exit(1)
    print(f"✅ Professional added to project")
    
    # Get allocation table to allocate hours
    print("\n4. Allocating hours...")
    alloc_table = requests.get(f"{BASE_URL}/projects/{project_id}/allocation_table")
    if alloc_table.status_code != 200:
        print(f"❌ Failed to get allocation table: {alloc_table.text}")
        sys.exit(1)
    alloc_data = alloc_table.json()
    
    # Update some weekly hours
    updates = []
    if len(alloc_data["allocations"]) > 0:
        allocation = alloc_data["allocations"][0]
        weekly_hours = allocation["weekly_hours"]
        
        # Set hours for first 2 weeks
        week_numbers = sorted(weekly_hours.keys(), key=int)[:2]
        for week_num_str in week_numbers:
            weekly_alloc = weekly_hours[week_num_str]
            available = weekly_alloc["available_hours"]
            hours_to_allocate = min(available * 0.8, 40.0)
            updates.append({
                "weekly_allocation_id": weekly_alloc["id"],
                "hours_allocated": hours_to_allocate
            })
    
    if updates:
        update_response = requests.put(f"{BASE_URL}/projects/{project_id}/allocations", json=updates)
        if update_response.status_code != 200:
            print(f"❌ Failed to update allocations: {update_response.text}")
            sys.exit(1)
        print(f"✅ Hours allocated for {len(updates)} weeks")
    
    # Export to Excel
    print("\n5. Exporting to Excel...")
    export_response = requests.get(f"{BASE_URL}/projects/{project_id}/export_excel")
    if export_response.status_code != 200:
        print(f"❌ Failed to export: {export_response.text}")
        sys.exit(1)
    
    if export_response.headers["content-type"] != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        print(f"❌ Wrong content type: {export_response.headers['content-type']}")
        sys.exit(1)
    
    if "attachment" not in export_response.headers["content-disposition"]:
        print(f"❌ Missing attachment header")
        sys.exit(1)
    
    print(f"✅ Excel file downloaded")
    
    # Load and verify Excel content
    print("\n6. Verifying Excel content...")
    excel_bytes = BytesIO(export_response.content)
    workbook = load_workbook(excel_bytes)
    
    # Verify sheet names
    sheet_names = workbook.sheetnames
    expected_sheets = ["Informações do Projeto", "Resumo Financeiro", "Tabela de Alocação"]
    for sheet in expected_sheets:
        if sheet not in sheet_names:
            print(f"❌ Missing sheet: {sheet}")
            sys.exit(1)
    print(f"✅ All expected sheets present")
    
    # Verify Project Info Sheet
    info_sheet = workbook["Informações do Projeto"]
    if info_sheet["A2"].value != "Nome do Projeto":
        print(f"❌ Wrong header in project info")
        sys.exit(1)
    if info_sheet["B2"].value != project_data["name"]:
        print(f"❌ Wrong project name in Excel")
        sys.exit(1)
    print(f"✅ Project info sheet verified")
    
    # Verify Financial Summary Sheet
    financial_sheet = workbook["Resumo Financeiro"]
    if financial_sheet["A1"].value != "Métrica":
        print(f"❌ Wrong header in financial summary")
        sys.exit(1)
    print(f"✅ Financial summary sheet verified")
    
    # Verify Allocation Table Sheet
    allocation_sheet = workbook["Tabela de Alocação"]
    if allocation_sheet["A1"].value != "ID":
        print(f"❌ Wrong header in allocation table")
        sys.exit(1)
    if allocation_sheet["B2"].value != professional["name"]:
        print(f"❌ Wrong professional name in allocation table")
        sys.exit(1)
    print(f"✅ Allocation table sheet verified")
    
    print("\n" + "="*60)
    print("✅ EXCEL EXPORT VERIFICATION COMPLETED SUCCESSFULLY!")
    print("="*60)

if __name__ == "__main__":
    try:
        verify_excel_export()
    except Exception as e:
        print(f"\n❌ Verification failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
